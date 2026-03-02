#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
from datetime import datetime
import json
import os
from pathlib import Path
from typing import Dict, List, Tuple
from zoneinfo import ZoneInfo

import boto3
from botocore.exceptions import BotoCoreError, ClientError
from openpyxl import load_workbook
import pymysql

if __package__ in (None, ""):  # pragma: no cover
    import sys

    sys.path.append(str(Path(__file__).resolve().parents[2]))

from scripts.flash_report_fill.checks import run_anomaly_checks
from scripts.flash_report_fill.excel_writer import (
    get_formula_cells,
    get_formula_protected_cells,
    is_updated_sheet_profile,
    write_flash_report_cells,
)
from scripts.flash_report_fill.reconciliation import (
    build_d5_discrepancy_records,
    build_reconciliation_records,
    metrics_records_to_csv_rows,
    reconciliation_records_to_csv_rows,
)
from scripts.flash_report_fill.sql import (
    ACTIVE_OCCUPANCY_STATUSES,
    COMPLETED_MOVEIN_STATUSES,
    COMPLETED_MOVEOUT_STATUSES,
    DEFAULT_MOVEIN_PREDICTION_COLUMN,
    DEFAULT_MOVEOUT_PREDICTION_COLUMN,
    INDIVIDUAL_CODES,
    PLANNED_MOVEIN_STATUSES,
    PLANNED_MOVEOUT_STATUSES,
    SUPPORTED_MOVEIN_PREDICTION_COLUMNS,
    SUPPORTED_MOVEOUT_PREDICTION_COLUMNS,
    build_metric_queries,
)
from scripts.flash_report_fill.types import FlashReportQueryConfig, MetricRecord, QuerySpec, WarningRecord


JST = ZoneInfo("Asia/Tokyo")
DEFAULT_TEMPLATE_PATH = "/Users/danielkang/Downloads/February Occupancy Flash Report_Template_GG追記20260224.xlsx"
DEFAULT_SECRET_ARN = "arn:aws:secretsmanager:ap-northeast-1:343881458651:secret:tokyobeta/prod/aurora/credentials-tlWiUd"
DEFAULT_DB_HOST = "tokyobeta-prod-aurora-cluster-public.cluster-cr46qo6y4bbb.ap-northeast-1.rds.amazonaws.com"
DEFAULT_DB_PORT = 3306
DEFAULT_DB_NAME = "tokyobeta"
DEFAULT_D5_BENCHMARK = 11271
DEFAULT_D5_TOLERANCE = 10

LEGACY_METRIC_TO_CELLS = {
    "d5_occupied_rooms": [("D5", "all", "2026-02")],
    "feb_completed_moveins": [("D11", "individual", "2026-02"), ("E11", "corporate", "2026-02")],
    "feb_planned_moveins": [("D12", "individual", "2026-02"), ("E12", "corporate", "2026-02")],
    "feb_completed_moveouts": [("D15", "individual", "2026-02"), ("E15", "corporate", "2026-02")],
    "feb_planned_moveouts": [("D16", "individual", "2026-02"), ("E16", "corporate", "2026-02")],
    "mar_completed_moveins": [("D13", "individual", "2026-03"), ("E13", "corporate", "2026-03")],
    "mar_planned_moveins": [("D14", "individual", "2026-03"), ("E14", "corporate", "2026-03")],
    "mar_planned_moveouts": [("D17", "individual", "2026-03"), ("E17", "corporate", "2026-03")],
}

UPDATED_METRIC_TO_CELLS = {
    **LEGACY_METRIC_TO_CELLS,
    "mar_completed_moveins": [],
    "mar_planned_moveins": [("D13", "individual", "2026-03"), ("E13", "corporate", "2026-03")],
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Fill February Occupancy Flash Report from Aurora")
    parser.add_argument("--template-path", default=DEFAULT_TEMPLATE_PATH, help="Excel template path")
    parser.add_argument("--sheet-name", default="Flash Report（2月）", help="Target sheet name")
    parser.add_argument("--output-dir", default=None, help="Output directory")
    parser.add_argument("--snapshot-start-jst", default="2026-02-01 00:00:00 JST")
    parser.add_argument("--snapshot-asof-jst", default="2026-02-28 05:00:00 JST")
    parser.add_argument(
        "--silver-pinpoint-asof-jst",
        default=None,
        help="Optional as-of timestamp for silver snapshot diagnostics; defaults to snapshot-asof date",
    )
    parser.add_argument("--feb-end-jst", default="2026-02-28 23:59:59 JST")
    parser.add_argument("--mar-start-jst", default="2026-03-01 00:00:00 JST")
    parser.add_argument("--mar-end-jst", default="2026-03-31 23:59:59 JST")
    parser.add_argument(
        "--movein-prediction-date-column",
        default=DEFAULT_MOVEIN_PREDICTION_COLUMN,
        choices=SUPPORTED_MOVEIN_PREDICTION_COLUMNS,
    )
    parser.add_argument(
        "--moveout-prediction-date-column",
        default=DEFAULT_MOVEOUT_PREDICTION_COLUMN,
        choices=SUPPORTED_MOVEOUT_PREDICTION_COLUMNS,
    )
    parser.add_argument("--d5-benchmark", type=int, default=DEFAULT_D5_BENCHMARK)
    parser.add_argument("--d5-tolerance", type=int, default=DEFAULT_D5_TOLERANCE)
    parser.add_argument("--aws-profile", default="gghouse")
    parser.add_argument("--aws-region", default="ap-northeast-1")
    parser.add_argument("--db-host", default=os.getenv("DB_HOST", DEFAULT_DB_HOST))
    parser.add_argument("--db-port", type=int, default=DEFAULT_DB_PORT)
    parser.add_argument("--db-name", default=DEFAULT_DB_NAME)
    parser.add_argument("--secret-arn", default=DEFAULT_SECRET_ARN)
    parser.add_argument("--db-user", default=None)
    parser.add_argument("--db-password", default=None)
    parser.add_argument("--emit-flags-csv", action="store_true", help="Export check A/B/C rows to CSV")
    parser.add_argument("--check-only", action="store_true", help="Run queries and output CSV without writing Excel")
    parser.add_argument("--flags-limit", type=int, default=300)
    return parser.parse_args()


def parse_jst_timestamp(value: str) -> datetime:
    cleaned = value.strip().replace("JST", "").strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S"):
        try:
            dt = datetime.strptime(cleaned, fmt)
            return dt.replace(tzinfo=JST)
        except ValueError:
            continue
    try:
        dt = datetime.fromisoformat(cleaned)
    except ValueError as exc:
        raise ValueError(f"Unsupported timestamp format: {value}") from exc
    if dt.tzinfo is None:
        return dt.replace(tzinfo=JST)
    return dt.astimezone(JST)


def resolve_db_credentials(args: argparse.Namespace) -> Tuple[str, str]:
    if args.db_user and args.db_password:
        return args.db_user, args.db_password

    session = boto3.Session(profile_name=args.aws_profile, region_name=args.aws_region)
    client = session.client("secretsmanager")

    try:
        response = client.get_secret_value(SecretId=args.secret_arn)
    except (BotoCoreError, ClientError) as exc:
        raise RuntimeError(f"Failed to load DB credentials from Secrets Manager: {exc}") from exc

    secret = json.loads(response["SecretString"])
    username = secret.get("username") or secret.get("user")
    password = secret.get("password")
    if not username or not password:
        raise RuntimeError("Secret did not contain username/password")
    return username, password


def open_connection(args: argparse.Namespace, user: str, password: str):
    return pymysql.connect(
        host=args.db_host,
        port=args.db_port,
        user=user,
        password=password,
        database=args.db_name,
        charset="utf8mb4",
        connect_timeout=30,
        read_timeout=120,
        write_timeout=120,
        cursorclass=pymysql.cursors.DictCursor,
    )


def execute_metric_query(cursor, spec: QuerySpec, params: Dict[str, object]) -> Dict[str, int] | int:
    cursor.execute(spec.sql, params)
    if spec.result_mode == "scalar":
        row = cursor.fetchone() or {}
        value = row.get(spec.value_column) if isinstance(row, dict) else row[0]
        return int(value or 0)

    rows = cursor.fetchall() or []
    split = {"individual": 0, "corporate": 0, "unknown": 0}
    for row in rows:
        tenant_type = str(row.get(spec.type_column, "unknown"))
        split[tenant_type] = int(row.get(spec.value_column, 0) or 0)
    return split


def write_csv(path: Path, rows: List[Dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        with path.open("w", newline="", encoding="utf-8") as handle:
            handle.write("")
        return
    fieldnames = list(rows[0].keys())
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def print_schema_mapping(query_config: FlashReportQueryConfig, sheet_name: str) -> None:
    print("=== SCHEMA MAPPING ===")
    print("source tables: staging.movings, staging.tenants, staging.rooms, staging.apartments")
    print("contract type: staging.movings.moving_agreement_type")
    print("move-in date (completed): staging.movings.original_movein_date")
    print(f"move-in date (prediction): staging.movings.{query_config.movein_prediction_date_column}")
    print("completed move-out date: staging.movings.moveout_date (date-priority)")
    print(f"planned move-out date: staging.movings.{query_config.moveout_prediction_date_column}")
    print(
        "status sets: "
        f"active={ACTIVE_OCCUPANCY_STATUSES}, "
        f"planned_movein={PLANNED_MOVEIN_STATUSES}, "
        f"completed_movein={COMPLETED_MOVEIN_STATUSES}, "
        f"completed_moveout={COMPLETED_MOVEOUT_STATUSES}, "
        f"planned_moveout={PLANNED_MOVEOUT_STATUSES}"
    )
    print("d5 mode: fact_aligned (strict archived)")
    print("d5 join/dedupe: movings.tenant_id -> tenants.id, tenant-room dedupe then room-priority dedupe")
    print(f"target sheet: {sheet_name}")
    print(f"contract split: individual={INDIVIDUAL_CODES}, corporate=(2, 3), unknown=excluded")
    print("======================")


def _window_for_metric(metric_id: str, params: Dict[str, object]) -> Tuple[str, str]:
    if metric_id.startswith("feb_"):
        return str(params["feb_start"]), str(params["feb_end"])
    if metric_id.startswith("mar_"):
        return str(params["mar_start"]), str(params["mar_end"])
    return str(params["snapshot_start"]), str(params["snapshot_start"])


def _resolve_sheet_name(template_path: Path, requested_sheet_name: str) -> str:
    workbook = load_workbook(template_path, read_only=True, data_only=False)
    try:
        if requested_sheet_name in workbook.sheetnames:
            return requested_sheet_name
        if requested_sheet_name == "Flash Report（2月）":
            for candidate in ("Flash Report（2月28日）", "Flash Report（3月～）"):
                if candidate in workbook.sheetnames:
                    return candidate
    finally:
        workbook.close()
    raise ValueError(f"Sheet not found in template: {requested_sheet_name}")


def _metric_to_cells_for_sheet(sheet_name: str) -> Dict[str, List[Tuple[str, str, str]]]:
    if is_updated_sheet_profile(sheet_name):
        return UPDATED_METRIC_TO_CELLS
    return LEGACY_METRIC_TO_CELLS


def main() -> int:
    args = parse_args()

    snapshot_start = parse_jst_timestamp(args.snapshot_start_jst)
    snapshot_asof = parse_jst_timestamp(args.snapshot_asof_jst)
    silver_pinpoint_asof = (
        parse_jst_timestamp(args.silver_pinpoint_asof_jst)
        if getattr(args, "silver_pinpoint_asof_jst", None)
        else snapshot_asof
    )
    feb_end = parse_jst_timestamp(args.feb_end_jst)
    mar_start = parse_jst_timestamp(args.mar_start_jst)
    mar_end = parse_jst_timestamp(args.mar_end_jst)
    feb_start = snapshot_start

    params: Dict[str, object] = {
        "snapshot_start": snapshot_start.strftime("%Y-%m-%d %H:%M:%S"),
        "snapshot_asof": snapshot_asof.strftime("%Y-%m-%d %H:%M:%S"),
        "feb_start": feb_start.strftime("%Y-%m-%d"),
        "feb_end": feb_end.strftime("%Y-%m-%d"),
        "mar_start": mar_start.strftime("%Y-%m-%d"),
        "mar_end": mar_end.strftime("%Y-%m-%d"),
    }

    template_path = Path(args.template_path).expanduser().resolve()
    output_dir = Path(args.output_dir).expanduser().resolve() if args.output_dir else template_path.parent
    now_jst = datetime.now(JST)
    ts_tag = now_jst.strftime("%Y%m%d_%H%M")

    output_xlsx = output_dir / f"February_Occupancy_FlashReport_filled_{ts_tag}.xlsx"
    metrics_csv = output_dir / f"flash_metrics_{ts_tag}.csv"
    recon_csv = output_dir / f"flash_reconciliation_{ts_tag}.csv"

    sheet_name = _resolve_sheet_name(template_path, args.sheet_name)
    query_config = FlashReportQueryConfig(
        movein_prediction_date_column=args.movein_prediction_date_column,
        moveout_prediction_date_column=args.moveout_prediction_date_column,
    )

    print_schema_mapping(query_config, sheet_name)
    print(f"template_path={template_path}")
    print(f"db_host={args.db_host}:{args.db_port}")
    print(f"snapshot_start={params['snapshot_start']}")
    print(f"snapshot_asof={params['snapshot_asof']}")
    print(f"silver_pinpoint_asof_date={silver_pinpoint_asof.date().isoformat()}")

    username, password = resolve_db_credentials(args)

    metric_records: List[MetricRecord] = []
    warnings: List[WarningRecord] = []
    cell_values: Dict[str, int] = {}
    metric_to_cells = _metric_to_cells_for_sheet(sheet_name)

    with open_connection(args, username, password) as connection:
        cursor = connection.cursor()

        queries = build_metric_queries(query_config)
        for metric_id, spec in queries.items():
            target_cells = metric_to_cells.get(metric_id, [])
            if not target_cells:
                continue

            result = execute_metric_query(cursor, spec, params)
            if spec.result_mode == "scalar":
                scalar_value = int(result)
                for cell, tenant_type, month in target_cells:
                    cell_values[cell] = scalar_value
                    window_start, window_end = _window_for_metric(metric_id, params)
                    metric_records.append(
                        MetricRecord(
                            metric_id=metric_id,
                            sheet=sheet_name,
                            cell=cell,
                            value=scalar_value,
                            month=month,
                            tenant_type=tenant_type,
                            window_start=window_start,
                            window_end=window_end,
                            asof_ts_jst=snapshot_asof.isoformat(),
                            source_layer=spec.source_layer,
                            query_tag=spec.query_tag,
                            query_sql=spec.sql,
                        )
                    )
                continue

            split_result = result
            unknown_count = int(split_result.get("unknown", 0))
            if unknown_count > 0:
                warnings.append(
                    WarningRecord(
                        code="WARN_UNKNOWN_CONTRACT_TYPE",
                        message=f"{metric_id} excluded unknown tenant_type records.",
                        count=unknown_count,
                    )
                )

            for cell, tenant_type, month in target_cells:
                value = int(split_result.get(tenant_type, 0))
                cell_values[cell] = value
                window_start, window_end = _window_for_metric(metric_id, params)
                metric_records.append(
                    MetricRecord(
                        metric_id=metric_id,
                        sheet=sheet_name,
                        cell=cell,
                        value=value,
                        month=month,
                        tenant_type=tenant_type,
                        window_start=window_start,
                        window_end=window_end,
                        asof_ts_jst=snapshot_asof.isoformat(),
                        source_layer=spec.source_layer,
                        query_tag=spec.query_tag,
                        query_sql=spec.sql,
                    )
                )

        if is_updated_sheet_profile(sheet_name):
            # Updated workbook includes a short-term corporate row; keep it fixed to zero by request.
            cell_values["D14"] = 0
            cell_values["E14"] = 0

        check_rows, check_warnings = run_anomaly_checks(cursor, params, limit=args.flags_limit)
        warnings.extend(check_warnings)

        recon_records = build_reconciliation_records(
            cursor=cursor,
            cell_values=cell_values,
            snapshot_start_date=snapshot_start.date(),
            snapshot_asof_date=snapshot_asof.date(),
            feb_start_date=feb_start.date(),
            feb_end_date=feb_end.date(),
            mar_start_date=mar_start.date(),
            mar_end_date=mar_end.date(),
            reconciliation_asof_date=silver_pinpoint_asof.date(),
            mar_planned_movein_cells=[
                metric_to_cells["mar_planned_moveins"][0][0],
                metric_to_cells["mar_planned_moveins"][1][0],
            ],
            movein_prediction_date_column=query_config.movein_prediction_date_column,
        )
        d5_recon_records, d5_warnings = build_d5_discrepancy_records(
            cursor=cursor,
            snapshot_start_ts=params["snapshot_start"],
            snapshot_start_date=snapshot_start.date(),
            benchmark_value=args.d5_benchmark,
            tolerance=args.d5_tolerance,
        )
        recon_records.extend(d5_recon_records)
        warnings.extend(d5_warnings)

    metrics_rows = metrics_records_to_csv_rows(metric_records)
    write_csv(metrics_csv, metrics_rows)
    write_csv(recon_csv, reconciliation_records_to_csv_rows(recon_records))

    if args.emit_flags_csv:
        for check_name, rows in check_rows.items():
            if not rows:
                continue
            flags_csv = output_dir / f"flags_{check_name}_{ts_tag}.csv"
            write_csv(flags_csv, rows)
            print(f"wrote {flags_csv}")

    if not args.check_only:
        formula_protected_cells = get_formula_protected_cells(sheet_name)
        before_formulas = get_formula_cells(template_path, sheet_name, sorted(formula_protected_cells))
        write_flash_report_cells(
            template_path=template_path,
            output_path=output_xlsx,
            sheet_name=sheet_name,
            values=cell_values,
        )
        after_formulas = get_formula_cells(output_xlsx, sheet_name, sorted(formula_protected_cells))
        if before_formulas != after_formulas:
            raise RuntimeError("Formula-protected cells changed unexpectedly.")
        print(f"wrote {output_xlsx}")

    print(f"wrote {metrics_csv}")
    print(f"wrote {recon_csv}")

    if warnings:
        print("=== WARNINGS ===")
        for warning in warnings:
            print(f"[{warning.code}] count={warning.count} {warning.message}")

    return 0


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())
