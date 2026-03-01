from __future__ import annotations

from pathlib import Path
from typing import Dict, Iterable

from openpyxl.cell.cell import MergedCell
from openpyxl import load_workbook


LEGACY_ALLOWED_INPUT_CELLS = {
    "D5",
    "D11",
    "E11",
    "D12",
    "E12",
    "D13",
    "E13",
    "D14",
    "E14",
    "D15",
    "E15",
    "D16",
    "E16",
    "D17",
    "E17",
}

LEGACY_FORMULA_PROTECTED_CELLS = {
    "E5",
    "H9",
    "H10",
    "H11",
    "H12",
    "H13",
    "H14",
    "H15",
    "D21",
    "E21",
    "D25",
    "E25",
    "D29",
    "E29",
}

UPDATED_FORMULA_PROTECTED_CELLS = {
    "E5",
    "I9",
    "I10",
    "I11",
    "I12",
    "I13",
    "I14",
    "I15",
    "D21",
    "E21",
    "D25",
    "E25",
    "D29",
    "E29",
}

UPDATED_SHEET_NAMES = {"Flash Report（2月28日）", "Flash Report（3月～）"}

# Backward-compatible defaults for callers/tests that still import these names.
ALLOWED_INPUT_CELLS = LEGACY_ALLOWED_INPUT_CELLS
FORMULA_PROTECTED_CELLS = LEGACY_FORMULA_PROTECTED_CELLS


def get_formula_protected_cells(sheet_name: str) -> set[str]:
    """Return formula-protected cells for the selected workbook sheet profile."""
    if is_updated_sheet_profile(sheet_name):
        return set(UPDATED_FORMULA_PROTECTED_CELLS)
    return set(LEGACY_FORMULA_PROTECTED_CELLS)


def get_allowed_input_cells(sheet_name: str) -> set[str]:
    """Return editable input cells for the selected workbook sheet profile."""
    if is_updated_sheet_profile(sheet_name):
        return set(LEGACY_ALLOWED_INPUT_CELLS)
    return set(LEGACY_ALLOWED_INPUT_CELLS)


def is_updated_sheet_profile(sheet_name: str) -> bool:
    """Detect updated sheet layout, including daily appended copies."""
    if sheet_name in UPDATED_SHEET_NAMES:
        return True
    return sheet_name.startswith("Flash Report（") and sheet_name.endswith("日）")


def write_flash_report_cells(
    template_path: Path,
    output_path: Path,
    sheet_name: str,
    values: Dict[str, int],
) -> None:
    """Write report input cells only, preserving formula-protected cells."""
    allowed_input_cells = get_allowed_input_cells(sheet_name)
    formula_protected_cells = get_formula_protected_cells(sheet_name)

    disallowed = sorted(set(values) - allowed_input_cells)
    if disallowed:
        raise ValueError(f"Attempted to write non-input cells: {', '.join(disallowed)}")

    protected_overwrites = sorted(set(values) & formula_protected_cells)
    if protected_overwrites:
        raise ValueError(f"Attempted to overwrite formula cells: {', '.join(protected_overwrites)}")

    workbook = load_workbook(template_path, data_only=False)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet not found: {sheet_name}")

    sheet = workbook[sheet_name]

    def _resolve_writable_cell(cell: str) -> str:
        target = sheet[cell]
        if not isinstance(target, MergedCell):
            return cell
        for merged_range in sheet.merged_cells.ranges:
            if cell in merged_range:
                return merged_range.start_cell.coordinate
        return cell

    resolved_values: Dict[str, int] = {}
    for cell, value in values.items():
        writable_cell = _resolve_writable_cell(cell)
        resolved_values[writable_cell] = resolved_values.get(writable_cell, 0) + int(value)

    protected_resolved_overwrites = sorted(set(resolved_values) & formula_protected_cells)
    if protected_resolved_overwrites:
        raise ValueError(
            "Attempted to overwrite formula cells: "
            + ", ".join(protected_resolved_overwrites)
        )

    for cell, value in resolved_values.items():
        sheet[cell] = int(value)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def get_formula_cells(path: Path, sheet_name: str, cells: Iterable[str]) -> Dict[str, str]:
    """Read formula values from the selected cells."""
    workbook = load_workbook(path, data_only=False)
    sheet = workbook[sheet_name]
    out: Dict[str, str] = {}
    for cell in cells:
        value = sheet[cell].value
        out[cell] = value if isinstance(value, str) else ""
    return out
