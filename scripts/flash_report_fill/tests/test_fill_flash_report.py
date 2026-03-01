from __future__ import annotations

from argparse import Namespace
from pathlib import Path

from botocore.exceptions import ClientError
from openpyxl import Workbook
import pytest

from scripts.flash_report_fill import fill_flash_report
from scripts.flash_report_fill.types import QuerySpec, ReconciliationRecord


class _MetricCursor:
    def __init__(self, row=None, rows=None):
        self._row = row or {}
        self._rows = rows or []

    def execute(self, sql, params):
        return None

    def fetchone(self):
        return self._row

    def fetchall(self):
        return self._rows


class _DummyConnection:
    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, tb):
        return None

    def cursor(self):
        return object()


class _CountCursor:
    def execute(self, sql):
        self._last_sql = sql

    def fetchone(self):
        return {"total_rooms": 123}


class _NoQueryCursor:
    def execute(self, sql):
        raise AssertionError("DB query should not be executed for fixed room denominator")

    def fetchone(self):
        return {"total_rooms": 99999}


def _create_template(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Flash Report（2月）"
    ws["D5"] = 100
    ws["E5"] = "=+D5/16109"
    wb.save(path)


def test_parse_jst_timestamp_accepts_jst_suffix() -> None:
    dt = fill_flash_report.parse_jst_timestamp("2026-02-01 00:00:00 JST")
    assert dt.year == 2026
    assert dt.month == 2
    assert dt.tzinfo is not None


def test_parse_jst_timestamp_rejects_invalid_input() -> None:
    with pytest.raises(ValueError):
        fill_flash_report.parse_jst_timestamp("not-a-timestamp")


def test_parse_args_defaults(monkeypatch) -> None:
    monkeypatch.setattr("sys.argv", ["fill_flash_report.py"])
    args = fill_flash_report.parse_args()
    assert args.db_host == fill_flash_report.DEFAULT_DB_HOST
    assert args.db_port == fill_flash_report.DEFAULT_DB_PORT
    assert args.sheet_name == "Flash Report（2月）"
    assert args.snapshot_asof_jst == "2026-02-28 05:00:00 JST"
    assert args.d5_mode == "fact_aligned"
    assert args.movein_prediction_date_column == "original_movein_date"
    assert args.moveout_prediction_date_column == "moveout_date"


def test_resolve_db_credentials_uses_cli_override() -> None:
    args = Namespace(db_user="user1", db_password="pw1")
    user, password = fill_flash_report.resolve_db_credentials(args)
    assert user == "user1"
    assert password == "pw1"


def test_resolve_db_credentials_reads_secret(monkeypatch) -> None:
    class _FakeClient:
        def get_secret_value(self, SecretId):
            return {"SecretString": '{"username": "u", "password": "p"}'}

    class _FakeSession:
        def client(self, name):
            assert name == "secretsmanager"
            return _FakeClient()

    monkeypatch.setattr(
        fill_flash_report.boto3,
        "Session",
        lambda profile_name, region_name: _FakeSession(),
    )

    args = Namespace(
        db_user=None,
        db_password=None,
        aws_profile="gghouse",
        aws_region="ap-northeast-1",
        secret_arn="arn:test",
    )
    user, password = fill_flash_report.resolve_db_credentials(args)
    assert user == "u"
    assert password == "p"


def test_resolve_db_credentials_raises_on_secret_error(monkeypatch) -> None:
    class _FakeClient:
        def get_secret_value(self, SecretId):
            raise ClientError({"Error": {"Code": "Boom", "Message": "x"}}, "GetSecretValue")

    class _FakeSession:
        def client(self, name):
            return _FakeClient()

    monkeypatch.setattr(
        fill_flash_report.boto3,
        "Session",
        lambda profile_name, region_name: _FakeSession(),
    )

    args = Namespace(
        db_user=None,
        db_password=None,
        aws_profile="gghouse",
        aws_region="ap-northeast-1",
        secret_arn="arn:test",
    )
    with pytest.raises(RuntimeError):
        fill_flash_report.resolve_db_credentials(args)


def test_execute_metric_query_scalar() -> None:
    cursor = _MetricCursor(row={"occupied_rooms": 77})
    spec = QuerySpec(
        metric_id="d5_occupied_rooms",
        sql="SELECT 77 AS occupied_rooms",
        source_layer="staging",
        result_mode="scalar",
        value_column="occupied_rooms",
    )
    value = fill_flash_report.execute_metric_query(cursor, spec, {})
    assert value == 77


def test_execute_metric_query_split() -> None:
    cursor = _MetricCursor(
        rows=[
            {"tenant_type": "individual", "cnt": 10},
            {"tenant_type": "corporate", "cnt": 3},
        ]
    )
    spec = QuerySpec(
        metric_id="feb_completed_moveins",
        sql="SELECT ...",
        source_layer="staging",
        result_mode="split",
        value_column="cnt",
    )
    value = fill_flash_report.execute_metric_query(cursor, spec, {})
    assert value["individual"] == 10
    assert value["corporate"] == 3
    assert value["unknown"] == 0


def test_open_connection_calls_pymysql(monkeypatch) -> None:
    called = {}

    def _fake_connect(**kwargs):
        called.update(kwargs)
        return "CONNECTION"

    monkeypatch.setattr(fill_flash_report.pymysql, "connect", _fake_connect)
    args = Namespace(db_host="127.0.0.1", db_port=3306, db_name="tokyobeta")
    conn = fill_flash_report.open_connection(args, "u", "p")
    assert conn == "CONNECTION"
    assert called["host"] == "127.0.0.1"


def test_window_and_total_rooms_helpers() -> None:
    params = {
        "feb_start": "2026-02-01",
        "feb_end": "2026-02-28",
        "mar_start": "2026-03-01",
        "mar_end": "2026-03-31",
        "snapshot_start": "2026-02-01 00:00:00",
    }
    assert fill_flash_report._window_for_metric("mar_planned_moveins", params) == ("2026-03-01", "2026-03-31")
    assert fill_flash_report._fetch_total_rooms(_CountCursor()) == 16109
    assert fill_flash_report._fetch_total_rooms(_NoQueryCursor()) == 16109


def test_main_check_only_writes_csv_outputs(tmp_path: Path, monkeypatch) -> None:
    template = tmp_path / "template.xlsx"
    _create_template(template)

    args = Namespace(
        template_path=str(template),
        sheet_name="Flash Report（2月）",
        output_dir=str(tmp_path),
        snapshot_start_jst="2026-02-01 00:00:00 JST",
        snapshot_asof_jst="2026-02-28 05:00:00 JST",
        feb_end_jst="2026-02-28 23:59:59 JST",
        mar_start_jst="2026-03-01 00:00:00 JST",
        mar_end_jst="2026-03-31 23:59:59 JST",
        d5_mode="fact_aligned",
        movein_prediction_date_column="original_movein_date",
        moveout_prediction_date_column="moveout_date",
        d5_benchmark=11271,
        d5_tolerance=10,
        aws_profile="gghouse",
        aws_region="ap-northeast-1",
        db_host="127.0.0.1",
        db_port=3306,
        db_name="tokyobeta",
        secret_arn="arn:test",
        db_user="user1",
        db_password="pw1",
        emit_flags_csv=True,
        check_only=True,
        flags_limit=10,
    )

    def _fake_execute_metric_query(cursor, spec, params):
        if spec.result_mode == "scalar":
            return 10000
        return {"individual": 10, "corporate": 5, "unknown": 1}

    monkeypatch.setattr(fill_flash_report, "parse_args", lambda: args)
    monkeypatch.setattr(fill_flash_report, "resolve_db_credentials", lambda _: ("u", "p"))
    monkeypatch.setattr(fill_flash_report, "open_connection", lambda *_args, **_kwargs: _DummyConnection())
    monkeypatch.setattr(fill_flash_report, "execute_metric_query", _fake_execute_metric_query)
    monkeypatch.setattr(fill_flash_report, "_fetch_total_rooms", lambda cursor: 16000)
    monkeypatch.setattr(
        fill_flash_report,
        "run_anomaly_checks",
        lambda cursor, params, limit: ({"double_active_rooms": [{"room_id": 1}], "tenant_switch_gaps": [], "same_tenant_multi_status": []}, []),
    )
    monkeypatch.setattr(
        fill_flash_report,
        "build_reconciliation_records",
        lambda **kwargs: [
            ReconciliationRecord(
                reconciliation_id="gold_feb_moveins_total",
                expected_value=1,
                reference_value=1,
                delta=0,
                reference_source="gold.occupancy_daily_metrics",
                note="ok",
                asof_date="2026-02-26",
            )
        ],
    )
    monkeypatch.setattr(fill_flash_report, "build_d5_discrepancy_records", lambda **kwargs: ([], []))

    rc = fill_flash_report.main()
    assert rc == 0
    assert list(tmp_path.glob("flash_metrics_*.csv"))
    assert list(tmp_path.glob("flash_reconciliation_*.csv"))
    assert list(tmp_path.glob("flags_double_active_rooms_*.csv"))


def test_main_write_mode_creates_filled_workbook(tmp_path: Path, monkeypatch) -> None:
    template = tmp_path / "template.xlsx"
    _create_template(template)

    args = Namespace(
        template_path=str(template),
        sheet_name="Flash Report（2月）",
        output_dir=str(tmp_path),
        snapshot_start_jst="2026-02-01 00:00:00 JST",
        snapshot_asof_jst="2026-02-28 05:00:00 JST",
        feb_end_jst="2026-02-28 23:59:59 JST",
        mar_start_jst="2026-03-01 00:00:00 JST",
        mar_end_jst="2026-03-31 23:59:59 JST",
        d5_mode="fact_aligned",
        movein_prediction_date_column="original_movein_date",
        moveout_prediction_date_column="moveout_date",
        d5_benchmark=11271,
        d5_tolerance=10,
        aws_profile="gghouse",
        aws_region="ap-northeast-1",
        db_host="127.0.0.1",
        db_port=3306,
        db_name="tokyobeta",
        secret_arn="arn:test",
        db_user="user1",
        db_password="pw1",
        emit_flags_csv=False,
        check_only=False,
        flags_limit=10,
    )

    def _fake_execute_metric_query(cursor, spec, params):
        if spec.result_mode == "scalar":
            return 10000
        return {"individual": 1, "corporate": 2, "unknown": 0}

    monkeypatch.setattr(fill_flash_report, "parse_args", lambda: args)
    monkeypatch.setattr(fill_flash_report, "resolve_db_credentials", lambda _: ("u", "p"))
    monkeypatch.setattr(fill_flash_report, "open_connection", lambda *_args, **_kwargs: _DummyConnection())
    monkeypatch.setattr(fill_flash_report, "execute_metric_query", _fake_execute_metric_query)
    monkeypatch.setattr(fill_flash_report, "_fetch_total_rooms", lambda cursor: 16109)
    monkeypatch.setattr(
        fill_flash_report,
        "run_anomaly_checks",
        lambda cursor, params, limit: ({"double_active_rooms": [], "tenant_switch_gaps": [], "same_tenant_multi_status": []}, []),
    )
    monkeypatch.setattr(fill_flash_report, "build_reconciliation_records", lambda **kwargs: [])
    monkeypatch.setattr(fill_flash_report, "build_d5_discrepancy_records", lambda **kwargs: ([], []))

    rc = fill_flash_report.main()
    assert rc == 0
    assert list(tmp_path.glob("February_Occupancy_FlashReport_filled_*.xlsx"))


def test_main_raises_if_formula_cells_change(tmp_path: Path, monkeypatch) -> None:
    template = tmp_path / "template.xlsx"
    _create_template(template)

    args = Namespace(
        template_path=str(template),
        sheet_name="Flash Report（2月）",
        output_dir=str(tmp_path),
        snapshot_start_jst="2026-02-01 00:00:00 JST",
        snapshot_asof_jst="2026-02-28 05:00:00 JST",
        feb_end_jst="2026-02-28 23:59:59 JST",
        mar_start_jst="2026-03-01 00:00:00 JST",
        mar_end_jst="2026-03-31 23:59:59 JST",
        d5_mode="fact_aligned",
        movein_prediction_date_column="original_movein_date",
        moveout_prediction_date_column="moveout_date",
        d5_benchmark=11271,
        d5_tolerance=10,
        aws_profile="gghouse",
        aws_region="ap-northeast-1",
        db_host="127.0.0.1",
        db_port=3306,
        db_name="tokyobeta",
        secret_arn="arn:test",
        db_user="user1",
        db_password="pw1",
        emit_flags_csv=False,
        check_only=False,
        flags_limit=10,
    )

    monkeypatch.setattr(fill_flash_report, "parse_args", lambda: args)
    monkeypatch.setattr(fill_flash_report, "resolve_db_credentials", lambda _: ("u", "p"))
    monkeypatch.setattr(fill_flash_report, "open_connection", lambda *_args, **_kwargs: _DummyConnection())
    monkeypatch.setattr(
        fill_flash_report,
        "execute_metric_query",
        lambda cursor, spec, params: 1 if spec.result_mode == "scalar" else {"individual": 1, "corporate": 1, "unknown": 0},
    )
    monkeypatch.setattr(fill_flash_report, "_fetch_total_rooms", lambda cursor: 16109)
    monkeypatch.setattr(
        fill_flash_report,
        "run_anomaly_checks",
        lambda cursor, params, limit: ({"double_active_rooms": [], "tenant_switch_gaps": [], "same_tenant_multi_status": []}, []),
    )
    monkeypatch.setattr(fill_flash_report, "build_reconciliation_records", lambda **kwargs: [])
    monkeypatch.setattr(fill_flash_report, "build_d5_discrepancy_records", lambda **kwargs: ([], []))
    monkeypatch.setattr(
        fill_flash_report,
        "get_formula_cells",
        lambda path, sheet_name, cells: {"E5": "A"} if str(path).endswith("template.xlsx") else {"E5": "B"},
    )

    with pytest.raises(RuntimeError):
        fill_flash_report.main()
