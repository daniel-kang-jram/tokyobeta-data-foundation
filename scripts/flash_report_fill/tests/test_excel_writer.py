from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
import pytest

from scripts.flash_report_fill.excel_writer import (
    get_formula_cells,
    get_formula_protected_cells,
    write_flash_report_cells,
)


def _create_template(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Flash Report（2月）"
    ws["D5"] = 100
    ws["E5"] = "=+D5/16109"
    ws["D11"] = 10
    ws["E11"] = 20
    wb.save(path)


def _create_updated_template(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Flash Report（2月28日）"
    ws["D5"] = 100
    ws["E5"] = "=+D5/16109"
    ws["D11"] = 10
    ws["E11"] = 20
    ws["I9"] = "=+D9/16109"
    ws["I10"] = "=+D10/16109"
    ws["I11"] = "=+D11/16109"
    ws["I12"] = "=+D12/16109"
    ws["I13"] = "=+D13/16109"
    ws["I14"] = "=+D14/16109"
    ws["I15"] = "=+D15/16109"
    wb.save(path)


def _create_updated_merged_template(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Flash Report（3月～）"
    ws["D5"] = 100
    ws["E5"] = "=+D5/16109"
    ws["I9"] = "=+D9/16109"
    ws["I10"] = "=+D10/16109"
    ws["I11"] = "=+D11/16109"
    ws["I12"] = "=+D12/16109"
    ws["I13"] = "=+D13/16109"
    ws["I14"] = "=+D14/16109"
    ws["I15"] = "=+D15/16109"
    ws.merge_cells("D11:E11")
    ws.merge_cells("D12:E12")
    wb.save(path)


def test_write_flash_report_cells_preserves_formula_cells(tmp_path: Path) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "out.xlsx"
    _create_template(template_path)

    write_flash_report_cells(
        template_path=template_path,
        output_path=output_path,
        sheet_name="Flash Report（2月）",
        values={"D5": 123, "D11": 99},
    )

    wb = load_workbook(output_path)
    ws = wb["Flash Report（2月）"]
    assert ws["D5"].value == 123
    assert ws["D11"].value == 99
    assert ws["E5"].value == "=+D5/16109"


def test_write_flash_report_cells_blocks_formula_cell_overwrite(tmp_path: Path) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "out.xlsx"
    _create_template(template_path)

    with pytest.raises(ValueError):
        write_flash_report_cells(
            template_path=template_path,
            output_path=output_path,
            sheet_name="Flash Report（2月）",
            values={"E5": 999},
        )


def test_write_flash_report_cells_blocks_non_input_cells(tmp_path: Path) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "out.xlsx"
    _create_template(template_path)

    with pytest.raises(ValueError):
        write_flash_report_cells(
            template_path=template_path,
            output_path=output_path,
            sheet_name="Flash Report（2月）",
            values={"A1": 123},
        )


def test_get_formula_cells_reads_formula_values(tmp_path: Path) -> None:
    template_path = tmp_path / "template.xlsx"
    _create_template(template_path)

    formulas = get_formula_cells(
        path=template_path,
        sheet_name="Flash Report（2月）",
        cells=("E5", "D5"),
    )
    assert formulas["E5"] == "=+D5/16109"
    assert formulas["D5"] == ""


def test_get_formula_protected_cells_supports_updated_sheet_profile() -> None:
    formula_cells = get_formula_protected_cells("Flash Report（2月28日）")
    assert "E5" in formula_cells
    assert "I9" in formula_cells
    assert "I15" in formula_cells


def test_write_flash_report_cells_preserves_updated_profile_formulas(tmp_path: Path) -> None:
    template_path = tmp_path / "updated_template.xlsx"
    output_path = tmp_path / "updated_out.xlsx"
    _create_updated_template(template_path)

    write_flash_report_cells(
        template_path=template_path,
        output_path=output_path,
        sheet_name="Flash Report（2月28日）",
        values={"D5": 111, "D11": 30, "E11": 9},
    )

    ws = load_workbook(output_path)["Flash Report（2月28日）"]
    assert ws["D5"].value == 111
    assert ws["D11"].value == 30
    assert ws["E11"].value == 9
    assert ws["I9"].value == "=+D9/16109"
    assert ws["I15"].value == "=+D15/16109"


def test_write_flash_report_cells_handles_merged_input_targets(tmp_path: Path) -> None:
    template_path = tmp_path / "updated_merged_template.xlsx"
    output_path = tmp_path / "updated_merged_out.xlsx"
    _create_updated_merged_template(template_path)

    write_flash_report_cells(
        template_path=template_path,
        output_path=output_path,
        sheet_name="Flash Report（3月～）",
        values={"D11": 30, "E11": 9, "D12": 5, "E12": 2},
    )

    ws = load_workbook(output_path)["Flash Report（3月～）"]
    assert ws["D11"].value == 39
    assert ws["D12"].value == 7
    assert ws["I11"].value == "=+D11/16109"
    assert ws["I12"].value == "=+D12/16109"
