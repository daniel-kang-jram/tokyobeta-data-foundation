#!/usr/bin/env python3
"""
Quick parse of rent roll Excel file - headers and sample only.
"""

import sys
import os

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system("pip3 install openpyxl --break-system-packages --quiet")
    import openpyxl

def quick_parse(file_path, max_rows=50):
    """Quick parse - headers and limited rows."""
    
    print(f"Reading file: {file_path}")
    print("="*80)
    
    # Load workbook (read-only for speed)
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    print(f"Sheets: {wb.sheetnames}\n")
    
    # Get first sheet
    sheet = wb.active
    print(f"Active sheet: {sheet.title}")
    print("="*80)
    
    # Read headers
    rows_iter = sheet.iter_rows(values_only=True)
    headers = next(rows_iter)
    headers = [h if h else f"col_{i}" for i, h in enumerate(headers)]
    
    print(f"\nHeaders ({len(headers)} columns):")
    for i, h in enumerate(headers[:50], 1):  # Show first 50
        print(f"{i:3d}. {h}")
    if len(headers) > 50:
        print(f"... and {len(headers) - 50} more columns")
    print("="*80)
    
    # Read sample data rows
    print(f"\nReading first {max_rows} rows...")
    data = []
    for i, row in enumerate(rows_iter):
        if i >= max_rows:
            break
        if any(row):  # Skip empty rows
            data.append(dict(zip(headers, row)))
    
    print(f"Read {len(data)} sample rows")
    print("="*80)
    
    # Display first row as sample
    print("\nSample data (first row):")
    for key, value in data[0].items():
        if value is not None:
            print(f"  {key}: {value}")
    
    # Quick analysis
    print("\n" + "="*80)
    print("QUICK ANALYSIS")
    print("="*80)
    
    # Identify key columns
    print("\nKey columns identified:")
    
    # Status columns
    status_cols = [h for h in headers if any(keyword in str(h) for keyword in ['ステータス', 'status', '状態', '入居状況'])]
    print(f"Status columns: {status_cols[:5]}")
    
    # Name columns
    name_cols = [h for h in headers if any(keyword in str(h) for keyword in ['氏名', 'name', '名前', 'テナント名'])]
    print(f"Name columns: {name_cols[:5]}")
    
    # Property columns
    property_cols = [h for h in headers if any(keyword in str(h) for keyword in ['物件', 'property', '建物', 'ビル'])]
    print(f"Property columns: {property_cols[:5]}")
    
    # Room columns
    room_cols = [h for h in headers if any(keyword in str(h) for keyword in ['部屋', 'room', '号室', 'unit'])]
    print(f"Room columns: {room_cols[:5]}")
    
    # Count non-null tenant names in sample
    if name_cols:
        col = name_cols[0]
        non_null = sum(1 for row in data if row.get(col) not in [None, '', ' '])
        print(f"\nSample: {non_null}/{len(data)} rows have tenant name in '{col}'")
    
    # Show sample status values
    if status_cols:
        col = status_cols[0]
        values = [row.get(col) for row in data if row.get(col) is not None]
        unique = set(values)
        print(f"\nSample status values in '{col}': {unique}")
    
    wb.close()
    return headers, data

if __name__ == "__main__":
    file_path = "data/RRデータ出力20251203.xlsx"
    
    if not os.path.exists(file_path):
        print(f"Error: File not found: {file_path}")
        sys.exit(1)
    
    headers, sample = quick_parse(file_path, max_rows=50)
    
    print("\n" + "="*80)
    print("To analyze full file, we need to:")
    print("1. Identify the correct status/occupancy column")
    print("2. Count rows where tenants are 'active' (入居中/occupied)")
    print("3. Compare with gold table active tenant count")
    print("="*80)
