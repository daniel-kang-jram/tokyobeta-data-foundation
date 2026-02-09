#!/usr/bin/env python3
"""
Parse rent roll Excel file and analyze active tenant count.
"""

import sys
import os
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system("pip3 install openpyxl --break-system-packages --quiet")
    import openpyxl

def parse_rent_roll(file_path):
    """Parse rent roll Excel file and return analysis."""
    
    print(f"Reading file: {file_path}")
    print("="*80)
    
    # Load workbook
    wb = openpyxl.load_workbook(file_path, data_only=True)
    print(f"Sheets: {wb.sheetnames}\n")
    
    # Get first sheet
    sheet = wb.active
    print(f"Active sheet: {sheet.title}")
    print(f"Dimensions: {sheet.dimensions}")
    print("="*80)
    
    # Read headers
    headers = []
    for cell in sheet[1]:
        headers.append(cell.value)
    
    print(f"\nHeaders ({len(headers)} columns):")
    for i, h in enumerate(headers, 1):
        if h:
            print(f"{i:3d}. {h}")
    print("="*80)
    
    # Read all data rows
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if any(row):  # Skip empty rows
            data.append(dict(zip(headers, row)))
    
    print(f"\nTotal rows: {len(data)}")
    print("="*80)
    
    # Display first 3 rows as sample
    print("\nSample data (first 3 rows):")
    for i, row in enumerate(data[:3], 1):
        print(f"\n--- Row {i} ---")
        for key, value in row.items():
            if value is not None and key:
                print(f"  {key}: {value}")
    
    # Analyze data
    print("\n" + "="*80)
    print("DATA ANALYSIS")
    print("="*80)
    
    # Try to identify status or occupancy columns
    print("\nLooking for status/occupancy indicators...")
    status_columns = [h for h in headers if h and any(keyword in str(h).lower() for keyword in ['status', 'ステータス', '状態', '入居', '退去', '空室', '稼働'])]
    if status_columns:
        print(f"Found potential status columns: {status_columns}")
        
        # Count by status
        for col in status_columns:
            values = [row.get(col) for row in data if row.get(col) is not None]
            unique_values = set(values)
            print(f"\n{col}:")
            print(f"  Unique values: {len(unique_values)}")
            value_counts = {}
            for v in values:
                value_counts[v] = value_counts.get(v, 0) + 1
            for v, count in sorted(value_counts.items(), key=lambda x: x[1], reverse=True)[:10]:
                print(f"    {v}: {count}")
    
    # Try to identify tenant name columns
    print("\n" + "-"*80)
    name_columns = [h for h in headers if h and any(keyword in str(h).lower() for keyword in ['name', '名前', '氏名', 'tenant', 'テナント'])]
    if name_columns:
        print(f"Found potential name columns: {name_columns}")
        for col in name_columns:
            non_null = sum(1 for row in data if row.get(col) is not None and row.get(col) != '')
            print(f"  {col}: {non_null} non-null values")
    
    # Try to identify property/room columns
    print("\n" + "-"*80)
    property_columns = [h for h in headers if h and any(keyword in str(h).lower() for keyword in ['property', '物件', 'room', '部屋', 'unit', 'building', 'ビル'])]
    if property_columns:
        print(f"Found potential property columns: {property_columns}")
        for col in property_columns[:3]:  # Show first 3
            values = [row.get(col) for row in data if row.get(col) is not None]
            unique = len(set(values))
            print(f"  {col}: {unique} unique values")
    
    # Calculate active tenant count estimation
    print("\n" + "="*80)
    print("ACTIVE TENANT COUNT ESTIMATION")
    print("="*80)
    
    # Method 1: Count non-empty tenant name rows
    if name_columns:
        active_by_name = sum(1 for row in data 
                            if any(row.get(col) not in [None, '', ' '] 
                                  for col in name_columns))
        print(f"Method 1 - Count by tenant name: {active_by_name}")
    
    # Method 2: Count by status if available
    if status_columns:
        # Try to identify "active" statuses
        for col in status_columns:
            values = [row.get(col) for row in data if row.get(col) is not None]
            # Common active indicators
            active_keywords = ['入居', '居住', 'occupied', 'active', '稼働']
            potentially_active = sum(1 for v in values 
                                    if any(keyword in str(v).lower() 
                                          for keyword in active_keywords))
            if potentially_active > 0:
                print(f"Method 2 - Count by status ({col}): {potentially_active} potentially active")
    
    print("\n" + "="*80)
    print(f"Total data rows in Excel: {len(data)}")
    print("="*80)
    
    return data, headers

if __name__ == "__main__":
    file_path = "data/RRデータ出力20251203.xlsx"
    
    if not os.path.exists(file_path):
        print(f"Error: File not found: {file_path}")
        sys.exit(1)
    
    data, headers = parse_rent_roll(file_path)
